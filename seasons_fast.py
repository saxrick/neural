from graph_fast import get_data
import sqlite3 as sq
from openpyxl import load_workbook

conn = sq.connect('Пациенты1.db')
cursor = conn.cursor()
sqlite_select_query = """SELECT * from ключи"""
cursor.execute(sqlite_select_query)
records = cursor.fetchall()


def get_data(row):
    data = f''
    return data


def season_graph(records):
    letters = 'BCDEFGHIJKLMNOPQRSTUVWXY'
    curr_wb = load_workbook('ГрафикиСезоны.xlsx')
    ws = curr_wb['гипотеза']
    conn = sq.connect('Пациенты1.db')
    cursor = conn.cursor()
    for i in records:
        ssq = f"""SELECT * from анализы WHERE (Ключ = '{i[1]}')"""
        cursor.execute(ssq)
        data = cursor.fetchall()
        c = 0
        for j in data:
            current = get_data(j)
            key = current['key']
            surname = current['surname']
            neu_lymf = round(current['neu'] / current['lymf'], 2)
            neu_cd3 = round(current['neu'] / current['cd3'], 2)
            neu_cd4 = round(current['neu'] / current['cd4'], 2)
            neu_cd8 = round(current['neu'] / current['cd8'], 2)

            lymf_cd19 = round(current['lymf'] / current['cd19'], 2)
            cd19_cd4 = round(current['cd19'] / current['cd4'], 2)
            cd19_cd8 = round(current['cd19'] / current['cd8'], 2)

            ws[f'{letters[c]}3'] = current['date']
            ws[f'{letters[c]}4'] = current['cd8']
            ws[f'{letters[c]}7'] = current['cd4']
            ws[f'{letters[c]}10'] = neu_lymf
            ws[f'{letters[c]}13'] = neu_cd3
            ws[f'{letters[c]}16'] = neu_cd8
            ws[f'{letters[c]}19'] = neu_cd4

            ws[f'{letters[c]}25'] = current['date']
            ws[f'{letters[c]}26'] = current['cd8']
            ws[f'{letters[c]}29'] = current['cd4']
            ws[f'{letters[c]}32'] = neu_lymf
            ws[f'{letters[c]}35'] = lymf_cd19
            ws[f'{letters[c]}38'] = cd19_cd4
            ws[f'{letters[c]}41'] = cd19_cd8
            c += 1
        curr_wb.save(f'графики по сезонам/{surname} {key}.xlsx')
        curr_wb.close()
season_graph(records)