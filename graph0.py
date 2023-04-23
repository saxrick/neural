from openpyxl import load_workbook
import sqlite3 as sq



def get_data():
    conn = sq.connect('Пациенты1.db')
    cursor = conn.cursor()
    sqlite_select_query = """SELECT * from анализы"""
    cursor.execute(sqlite_select_query)
    records = cursor.fetchall()
    data = {
        'lymf': xl_list['C10'].value,
        'neu': xl_list['C14'].value,
        'cd3': xl_list['C29'].value,
        'cd4': xl_list['C31'].value,
        'cd8': xl_list['C33'].value,
        'cd19': xl_list['C27'].value,
        'date': str(xl_list['C5'].value)[0:10].replace('-', '.')
    }
    return data


def grafik():
    file = 'Результаты_Анализов.xlsx'
    wb = load_workbook(file)
    curr_wb = load_workbook('Графики.xlsx')
    for sheet in wb.worksheets[2:]:
        ws = curr_wb['grafik']
        a = str(sheet)[12:-2]
        xl_list = wb[a]
        data = get_data()

        neu_lymf = round(data['neu'] / data['lymf'], 2)
        neu_cd3 = round(data['neu'] / data['cd3'], 2)
        neu_cd4 = round(data['neu'] / data['cd4'], 2)
        neu_cd8 = round(data['neu'] / data['cd8'], 2)

        lymf_cd19 = round(data['lymf'] / data['cd19'], 2)
        cd19_cd4 = round(data['cd19'] / data['cd4'], 2)
        cd19_cd8 = round(data['cd19'] / data['cd8'], 2)

        ws['F14'] = neu_lymf
        ws['F15'] = neu_cd3
        ws['F16'] = neu_cd4
        ws['F17'] = neu_cd8

        ws['F23'] = neu_lymf
        ws['F24'] = lymf_cd19
        ws['F25'] = cd19_cd4
        ws['F26'] = cd19_cd8

        ws1 = curr_wb['grafik1']

        ws1['C3'] = data['date']
        ws1['C4'] = data['cd8']
        ws1['C7'] = data['cd4']
        ws1['C10'] = round(data['neu'] / data['lymf'], 2)
        ws1['C13'] = round(data['neu'] / data['cd3'], 2)
        ws1['C16'] = round(data['neu'] / data['cd8'], 2)
        ws1['C19'] = round(data['neu'] / data['cd4'], 2)

        ws1['C25'] = data['date']
        ws1['C26'] = data['cd8']
        ws1['C29'] = data['cd4']
        ws1['C32'] = round(data['neu'] / data['lymf'], 2)
        ws1['C35'] = round(data['neu'] / data['cd19'], 2)
        ws1['C38'] = round(data['neu'] / data['cd4'], 2)
        ws1['C41'] = round(data['neu'] / data['cd8'], 2)
        curr_wb.save(f'графики пациентов/{a}.xlsx')
        curr_wb.close()

grafik()



