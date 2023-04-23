from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook
from tkinter.ttk import Combobox
import sqlite3 as sq
from openpyxl import load_workbook
import random
import string


# создание окна
ws = Tk()
ws.title('Ввод значений')
ws.geometry('800x620')
ws["bg"] = "gray80"

def welcomeMessage():
    # конвертация введенных данных
    lymf = float(name_LYMF.get())
    neu = float(name_NEU.get())
    cd3 = float(name_CD3.get())
    cd4 = float(name_CD4.get())
    cd8 = float(name_CD8.get())
    cd19 = float(name_CD19.get())
    wbc = float(name_WBC.get())
    cd3ifn_ct = float(name_CD3IFN_CT.get())
    cd3ifn_co = float(name_CD3IFN_CO.get())
    mon = float(name_MON.get())
    cd3inf_ct = float(name_CD3INF_CT.get())
    age = float(name_age.get())
    cd3ing_co = float(name_CD3INF_CO.get())
    eos = float(name_EOS .get())
    rato = float(name_ratio.get())
    cd3il2_ct = float(name_CD3IL2_CT.get())
    bas = float(name_BAS.get())
    nk_ob = float(name_NK_Ob.get())
    cd3il2_co = float(name_CD3IL2_CO.get())
    gens1 = float(name_gens1.get())
    hgb = float(name_HGB.get())
    nk_cit = float(name_NK_Cit.get())
    fno = float(name_FNO.get())
    gens2 = float(name_2gens.get())
    plt = float(name_PLT.get())
    cit = float(name_CIK.get())
    gens3 = float(name_3gens.get())
    hci_co = float(name_HCI_CO.get())
    hci_ct = float(name_HCI_CT.get())

    def generate_alphanum_random_string(length):
        letters_and_digits = string.ascii_letters + string.digits
        rand_string = ''.join(random.sample(letters_and_digits, length))
        return rand_string

    conn = sq.connect('опездалы.db')
    cur = conn.cursor()
    cur.executescript(f'''PRAGMA foreign_keys=on;
        CREATE TABLE IF NOT EXISTS анализы(
        userid INT PRIMARY KEY,
        Ключ TEXT,
        ФИО TEXT,
        Пол TEXT,   
        Дата_Анализа TEXT,
        Возраст TEXT,
        Диагноз_основной TEXT,
        Диагноз_сопутствующий TEXT,
        Гены1 TEXT,
        Гены2 TEXT,
        Гены3 TEXT,  
        Сезон TEXT,
        Лейкоциты_WBC TEXT,
        Лимфоциты_LYMF TEXT,
        Моноциты_MON TEXT,
        Нейтрофилы_NEU TEXT,
        Гемоглобин_HGB TEXT,
        Тромбоциты_PLT TEXT,
        Эозинофилы_EOS TEXT,
        Базофилы_BAS TEXT,
        Общие_B_лимфоциты TEXT,
        Общие_T_лимфоциты TEXT,
        Т_хелперы TEXT,
        Т_цитотоксические лимфоциты TEXT,
        Соотношение_CD3_CD4_CD3_CD8 TEXT,
        NK_клетки_цитолитические TEXT,
        Общие_NK_клетки TEXT,
        Циркулирующие_имунные_комплексы TEXT,
        НСТ_тест_спонтанный TEXT,
        НСТ_тест_стимулированный TEXT,  
        CD3_IFNy_стимулированый TEXT,
        CD3_IFNy_спонтанный TEXT,
        CD3_TNFa_стимулированный TEXT,
        CD3_TNFa_спонтанный TEXT,
        CD3_IL2_стимулированный TEXT,
        CD3_IL2_спонтанный TEXT,
        ФНО TEXT)''')

    conn.commit()

    file = 'результаты_анализов_пример_расчета_автоматические_графики_для_урфу.xlsx'
    wb = load_workbook(file)
    db_list = []
    name_list = []
    c = 0
    for sheet in wb.worksheets[2:]:
        a = str(sheet)[12:-2]
        xl_list = wb[a]
        date1 = str(xl_list['C5'].value).replace('-', '.')
        date1 = date1[:10]
        name = xl_list['C3'].value
        name_list.append(name)

        db_list.append((c, '', xl_list['C3'].value, 'NULL', date1, xl_list['C4'].value, 'NULL', 'NULL', 'NULL', 'NULL',
                        'NULL', xl_list['A6'].value,

                        xl_list['C9'].value, xl_list['C10'].value, xl_list['C12'].value, xl_list['C14'].value,
                        xl_list['C16'].value,
                        xl_list['C18'].value, 'NULL', 'NULL',

                        xl_list['C27'].value, xl_list['C29'].value, xl_list['C31'].value,
                        xl_list['C33'].value, 'NULL', xl_list['C37'].value, xl_list['C38'].value, xl_list['C24'].value,
                        xl_list['C51'].value, xl_list['C57'].value,

                        xl_list['C46'].value, xl_list['C47'].value, xl_list['C41'].value,
                        xl_list['C40'].value, xl_list['C42'].value, xl_list['C43'].value, 'NULL'))
        c += 1

    temp = []
    for x in name_list:
        if x not in temp:
            temp.append(x)

    res_list = []
    for i in temp:
        curr = (i, generate_alphanum_random_string(5))
        res_list.append(curr)
    print(res_list)
    cur.executemany("INSERT INTO анализы VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);",db_list)
    conn.commit()

    cur.executescript(f'''PRAGMA foreign_keys=on;
    CREATE TABLE IF NOT EXISTS keys(фамилия TEXT PRIMARY KEY,
    ключ TEXT)''')

    cur.executemany("INSERT INTO keys VALUES(?, ?)", res_list)
    conn.commit()
    key = generate_alphanum_random_string(5)
    count = '20'
    data_list = [(count, key, str(Surname.get()), str(name_ManWom.get()), str(name_dateAnal.get()), str(name_age.get()), str(name_DiagnosBasic.get()), str(name_DiagnosRelated.get()), str(name_gens1.get()), str(name_2gens.get()), str(name_3gens.get()), str(name_Season.get()),str(name_WBC.get()), str(name_LYMF.get()), str(name_MON.get()), str(name_NEU.get()), str(name_HGB.get()),str(name_PLT.get()), str(name_EOS.get()), str(name_BAS.get()), str(name_CD3.get()), str(name_CD19.get()), str(name_CD4.get()), str(name_CD8.get()), str(name_ratio.get()), str(name_NK_Ob.get()), str(name_NK_Cit.get()), str(name_CIK.get()), str(name_HCI_CO.get()), str(name_HCI_CT.get()), str(name_CD3IFN_CO.get()), str(name_CD3IFN_CT.get()), str(name_CD3INF_CO.get()),str(name_CD3INF_CT.get()), str(name_CD3IL2_CO.get()), str(name_CD3IL2_CT.get()), str(name_FNO.get()))]
    for i in data_list:
        for j in i:
            print(type(j), '   ', j)
    cur.executemany("INSERT INTO анализы VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", data_list)
    conn.commit()

    # создание графиков
    wb = Workbook()
    ws = wb.active
    file = 'результаты_анализов_пример_расчета_автоматические_графики_для_урфу.xlsx'
    wb = load_workbook(file)
    curr_wb = load_workbook('grafik.xlsx')
    ws = curr_wb['grafik']

    for sheet in wb.worksheets[2:]:
        a = str(sheet)[12:-2]
        xl_list = wb[a]
        name = a

        ws['F14'] = neu / lymf
        ws['F15'] = neu / cd3
        ws['F16'] = neu / cd4
        ws['F17'] = neu / cd8

        ws['F23'] = neu / lymf
        ws['F24'] = lymf / cd19
        ws['F25'] = cd19 / cd4
        ws['F26'] = cd19 / cd8
        curr_wb.save('./results.xlsx')
        curr_wb.close()
    return messagebox.showinfo('Вывод', "Данные сохраненны")

#создание объектов
Label(ws, bg='gray81', font="TkDefaultFont 11",text="Выберите из списка или создайте новую").place(x=20, y=10)

combo = Combobox(ws)
combo['values'] = ("none", "Тиккоева", "Чеснокова", "Баринов", "Зуева", "Корзухина",
                   "Абрамов", "Бородина", "Немцев", "Зубов", "Тарасова", "Дылдин", "Сорокина")
combo.current(1)
combo.place(x=22, y=35)

btn = Button(ws, text="    НОВЫЙ    ", command=welcomeMessage)
btn.place(x=200, y=33)

btn = Button(ws, text="    Ввод    ", command=welcomeMessage)
btn.place(x=700, y=20)

#0 заглавки

Label(ws, bg='gray81', font="TkDefaultFont 11",text="Персональные данные").place(x=25, y=60)
Label(ws, bg='gray81',font="TkDefaultFont 11", text="Резульаты гемоталогического\nисследования").place(x=205, y=60)
Label(ws, bg='gray81',font="TkDefaultFont 11", text="Имунный статус").place(x=440, y=63)
Label(ws, bg='gray81',font="TkDefaultFont 11", text="Цитокиновый статус").place(x=625, y=62)

#1

Label(ws, bg='gray81', text="Фамилия").place(x=70, y=100)
Surname = Entry(ws)
Surname.place(x=37.5, y=120)

Label(ws, bg='gray81', text="Лейкоциты(WBC)").place(x=252, y=100)
name_WBC = Entry(ws)
name_WBC.place(x=237.5, y=120)

Label(ws, bg='gray81', text="Общие Т-Лимфоциты").place(x=440, y=100)
name_CD3 = Entry(ws)
name_CD3.place(x=437.5, y=120)

Label(ws, bg='gray81', text="CD3+IFN+(стимулированный)").place(x=620, y=100)
name_CD3IFN_CT = Entry(ws)
name_CD3IFN_CT.place(x=637.5, y=120)

#2

Label(ws, bg='gray81', text="Пол(м/ж)").place(x=70, y=150)
name_ManWom = Entry(ws)
name_ManWom.place(x=37.5, y=170)

Label(ws, bg='gray81', text="Леймфоциты(LYMF)").place(x=247, y=150)
name_LYMF = Entry(ws)
name_LYMF.place(x=237.5, y=170)

Label(ws, bg='gray81', text="Общие B-Лимфоциты").place(x=440, y=150)
name_CD19 = Entry(ws)
name_CD19.place(x=437.5, y=170)

Label(ws, bg='gray81', text="CD3+IFN+(спонтанный)").place(x=635, y=150)
name_CD3IFN_CO = Entry(ws)
name_CD3IFN_CO.place(x=637.5, y=170)

#3

Label(ws, bg='gray81', text="Дата анализа").place(x=67, y=200)
name_dateAnal = Entry(ws)
name_dateAnal.place(x=37.5, y=220)

Label(ws, bg='gray81', text="Моноциты(MON)").place(x=252, y=200)
name_MON = Entry(ws)
name_MON.place(x=237.5, y=220)

Label(ws, bg='gray81', text="T-хелперы").place(x=470, y=200)
name_CD4 = Entry(ws)
name_CD4.place(x=437.5, y=220)

Label(ws, bg='gray81', text="CD3+TNFa+(стимулированный)").place(x=610, y=200)
name_CD3INF_CT = Entry(ws)
name_CD3INF_CT.place(x=637.5, y=220)

#4

Label(ws, bg='gray81', text="Возраст пациента").place(x=52, y=250)
name_age = Entry(ws)
name_age.place(x=37.5, y=270)

Label(ws, bg='gray81', text="Нейтрофилы(NEU)").place(x=247, y=250)
name_NEU = Entry(ws)
name_NEU.place(x=237.5, y=270)

Label(ws, bg='gray81', text="Т-цитотоксические лимфоциты").place(x=415, y=250)
name_CD8 = Entry(ws)
name_CD8.place(x=437.5, y=270)

Label(ws, bg='gray81', text="CD3+TNFa+(спонтанный)").place(x=625, y=250)
name_CD3INF_CO = Entry(ws)
name_CD3INF_CO.place(x=637.5, y=270)

#5

Label(ws, bg='gray81', text="Диагноз оснвной").place(x=51, y=300)
name_DiagnosBasic = Entry(ws)
name_DiagnosBasic.place(x=37.5, y=320)

Label(ws, bg='gray81', text="Эозофилы(EOS)").place(x=257, y=300)
name_EOS = Entry(ws)
name_EOS.place(x=237.5, y=320)

Label(ws, bg='gray81', text="Соотношение CD3+CD4+/CD3+CD8+").place(x=405, y=300)
name_ratio = Entry(ws)
name_ratio.place(x=437.5, y=320)

Label(ws, bg='gray81', text="CD3+IL2+(стимулированный)").place(x=620, y=300)
name_CD3IL2_CT = Entry(ws)
name_CD3IL2_CT.place(x=637.5, y=320)

#6

Label(ws, bg='gray81', text="Дигноз сопустствующий").place(x=32, y=350)
name_DiagnosRelated = Entry(ws)
name_DiagnosRelated.place(x=37.5, y=370)

Label(ws, bg='gray81', text="Базофилы(BAS)").place(x=257, y=350)
name_BAS = Entry(ws)
name_BAS.place(x=237.5, y=370)

Label(ws, bg='gray81', text="Общие NK-клетки").place(x=447, y=350)
name_NK_Ob = Entry(ws)
name_NK_Ob.place(x=437.5, y=370)

Label(ws, bg='gray81', text="CD3+IL2+(спонтанный)").place(x=635, y=350)
name_CD3IL2_CO = Entry(ws)
name_CD3IL2_CO.place(x=637.5, y=370)

#7

Label(ws, bg='gray81', text="1.Гены").place(x=80, y=400)
name_gens1 = Entry(ws)
name_gens1.place(x=37.5, y=420)

Label(ws, bg='gray81', text="Гемоглобин(HGB)").place(x=252, y=400)
name_HGB = Entry(ws)
name_HGB.place(x=237.5, y=420)

Label(ws, bg='gray81', text="NK-клетки цитолитические").place(x=425, y=400)
name_NK_Cit = Entry(ws)
name_NK_Cit.place(x=437.5, y=420)

Label(ws, bg='gray81', text="ФНО").place(x=685, y=400)
name_FNO = Entry(ws)
name_FNO.place(x=637.5, y=420)

#8

Label(ws, bg='gray81', text="2.Гены").place(x=80, y=450)
name_2gens = Entry(ws)
name_2gens.place(x=37.5, y=470)

Label(ws, bg='gray81', text="Тромбоциты(PLT)").place(x=252, y=450)
name_PLT = Entry(ws)
name_PLT.place(x=237.5, y=470)

Label(ws, bg='gray81', text="Цируклирующие имунные комплексы").place(x=395, y=450)
name_CIK = Entry(ws)
name_CIK.place(x=437.5, y=470)

#9

Label(ws, bg='gray81', text="3.Гены").place(x=80, y=500)
name_3gens = Entry(ws)
name_3gens.place(x=37.5, y=520)

Label(ws, bg='gray81', text="HCI-тест(спонтанный)").place(x=440, y=500)
name_HCI_CO = Entry(ws)
name_HCI_CO.place(x=437.5, y=520)

#10

Label(ws, bg='gray81', text="Сезон").place(x=82, y=550)
name_Season = Entry(ws)
name_Season.place(x=37.5, y=570)

Label(ws, bg='gray81', text="HCI-тест(стимулированный)").place(x=422, y=550)
name_HCI_CT = Entry(ws)
name_HCI_CT.place(x=437.5, y=570)

ws.mainloop()
wb = Workbook()
ws = wb.active
