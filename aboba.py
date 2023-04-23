from openpyxl import Workbook, load_workbook
import sqlite3 as sq

conn = sq.connect('Пациенты.db')
cursor = conn.cursor()
sqlite_select_query = """SELECT * from анализы"""
cursor.execute(sqlite_select_query)
records = cursor.fetchall()
surnames = []
cd3_cd4 = []
cd3_cd8 = []

for row in records:
    surnames.append(row[2])
    cd3_cd4.append((float(row[21]) + float(row[22])) / 2)
    cd3_cd8.append((float(row[22]) + float(row[23])) / 2)

sur_cd4 = dict(zip(surnames, cd3_cd4))
sur_cd8 = dict(zip(surnames, cd3_cd8))
sur_cd4_sorted = dict(sorted(sur_cd4.items(), key=lambda item: item[1]))
sur_cd8_sorted = dict(sorted(sur_cd8.items(), key=lambda item: item[1]))

wb = Workbook()
curr_wb = load_workbook('new_graph.xlsx')
ws = curr_wb['grafik']
i = 2
for key, value in sur_cd4_sorted.items():
    ws[f'A{i}'] = key
    ws[f'B{i}'] = value
    i += 1
i = 2
for key, value in sur_cd8_sorted.items():
    ws[f'C{i}'] = key
    ws[f'D{i}'] = value
    i += 1
curr_wb.save('new_graph.xlsx')
curr_wb.close()