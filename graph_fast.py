import sqlite3 as sq
from openpyxl import load_workbook

conn = sq.connect('Пациенты1.db')
cursor = conn.cursor()
sqlite_select_query = """SELECT * from анализы"""
cursor.execute(sqlite_select_query)
records = cursor.fetchall()


def get_data(row):
    data = {
        'key': row[1],
        'surname': row[2],
        'lymf': float(row[13]),
        'neu': float(row[15]),
        'cd3': float(row[21]),
        'cd4': float(row[22]),
        'cd8': float(row[23]),
        'cd19': float(row[20]),
        'date': row[4][0:10].replace('-', '.'),

    }
    return data


def data_graph3(row):
    if row[30] is None or row[31] is None or row[32] is None or row[33] is None or row[34] is None or row[35] is None:
        data = {
            'CD3+IFNy+st': 1,
            'CD3+IFNy+sp': 1,
            'CD3+TNFa+st': 1,
            'CD3+TNFa+sp': 1,
            'CD3+IL2+st': 1,
            'CD3+IL2+sp': 1,
        }
    else:
        data = {
            'CD3+IFNy+st': float(row[30]),
            'CD3+IFNy+sp': float(row[31]),
            'CD3+TNFa+st': float(row[32]),
            'CD3+TNFa+sp': float(row[33]),
            'CD3+IL2+st': float(row[34]),
            'CD3+IL2+sp': float(row[35]),
        }
    return data


def grafik(records):
    curr_wb = load_workbook('Графики.xlsx')
    ws = curr_wb['grafik']
    for i in records:
        data = get_data(i)
        data1 = data_graph3(i)
        surname = data['surname']
        date = data['date']

        neu_lymf = round(data['neu'] / data['lymf'], 2)
        neu_cd3 = round(data['neu'] / data['cd3'], 2)
        neu_cd4 = round(data['neu'] / data['cd4'], 2)
        neu_cd8 = round(data['neu'] / data['cd8'], 2)

        lymf_cd19 = round(data['lymf'] / data['cd19'], 2)
        cd19_cd4 = round(data['cd19'] / data['cd4'], 2)
        cd19_cd8 = round(data['cd19'] / data['cd8'], 2)

        CD3_IFNy = round(data1['CD3+IFNy+st'] / data1['CD3+IFNy+sp'], 2)
        CD3_TNFa = round(data1['CD3+TNFa+st'] / data1['CD3+TNFa+sp'], 2)
        CD3_IL2 = round(data1['CD3+IL2+st'] / data1['CD3+IL2+sp'], 2)

        ws['F14'] = neu_lymf
        ws['F15'] = neu_cd3
        ws['F16'] = neu_cd4
        ws['F17'] = neu_cd8

        ws['F23'] = neu_lymf
        ws['F24'] = lymf_cd19
        ws['F25'] = cd19_cd4
        ws['F26'] = cd19_cd8

        ws['M22'] = CD3_IFNy
        ws['M23'] = CD3_TNFa
        ws['M24'] = CD3_IL2

        curr_wb.save(f'графики пациентов/{surname} {date}.xlsx')
        curr_wb.close()

grafik(records)